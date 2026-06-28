"""
excel_generator.py
─────────────────────────
Builds a formatted .xlsx report from scraped case results.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import (
    get_column_letter,
)


# ─────────────────────────────────────────────────────────────────────────────
# COLUMNS
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS = [

    ("CNR Number",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "cnr_number",
            r.get(
                "cnr",
                "",
            ),
        )),

    ("Case Type",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "case_type",
            "",
        )),

    ("Case Stage / Status",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "case_stage",
            "",
        )),

    ("Court Name",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "court_name",
            "",
        )),

    ("State",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "state",
            "",
        )),

    ("Filing Date",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "filing_date",
            "",
        )),

    ("Latest Hearing Date",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "latest_hearing_date",
            "",
        )),

    ("Next Hearing Date",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "next_hearing_date",
            "",
        )),

    ("Lawyer",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "lawyer",
            "",
        )),

    ("Opposite Party",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "opposite_party",
            "",
        )),

    ("OP Lawyer",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "op_lawyer",
            "",
        )),

    ("Acts & Sections",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "acts",
            "",
        )),

    ("Total Hearings",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "hearing_count",
            "",
        )),

    ("Consec. Adjournments",
     lambda r:
        r.get(
            "data",
            {}
        ).get(
            "consecutive_adjournments",
            "",
        )),

    ("PDF Pages",
     lambda r:
        r.get(
            "pdf_pages",
            "",
        )),

    ("PDF Size KB",
     lambda r:
        r.get(
            "pdf_size_kb",
            "",
        )),

    ("Processing Time (s)",
     lambda r:
        r.get(
            "processing_time",
            "",
        )),

    ("AI Outcome",
     lambda r:
        r.get(
            "ai_outcome",
            "",
        )),

    ("AI Next Action",
     lambda r:
        r.get(
            "ai_next_action",
            "",
        )),

    ("AI Key Parties",
     lambda r:
        r.get(
            "ai_key_parties",
            "",
        )),

    ("Scrape Status",
     lambda r:
        (
            "✓ Done"
            if r.get(
                "status"
            ) == "done"
            else
            f"✗ Failed: "
            f"{r.get('error','')[:60]}"
        )),
]


# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────

C_TITLE_BG = "1F4E79"
C_HEADER_BG = "2E75B6"
C_HEADER_FG = "FFFFFF"
C_ALT_ROW = "D6E4F0"
C_FAIL_ROW = "FFE4E4"
C_SUMMARY_BG = "EEF4FB"


def _thin_border():

    s = Side(
        style="thin",
        color="BBBBBB",
    )

    return Border(
        left=s,
        right=s,
        top=s,
        bottom=s,
    )


def _set_col_width(
    ws,
    col_idx,
    value,
):

    letter = (
        get_column_letter(
            col_idx
        )
    )

    current = (
        ws.column_dimensions[
            letter
        ].width
        or 10
    )

    needed = min(
        max(
            len(
                str(value)
            ) + 4,
            14,
        ),
        60,
    )

    if needed > current:
        ws.column_dimensions[
            letter
        ].width = needed


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(
    cases,
    user_name,
    output_dir="data",
):

    os.makedirs(
        output_dir,
        exist_ok=True,
    )

    stamp = (
        datetime.now()
        .strftime(
            "%Y%m%d_%H%M%S"
        )
    )

    safe_name = (
        user_name
        .replace(
            " ",
            "_",
        )
        .replace(
            "/",
            "-",
        )
    )

    filepath = os.path.join(
        output_dir,
        f"eCourts_"
        f"{safe_name}_"
        f"{stamp}.xlsx",
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "Case Report"

    num_cols = len(
        COLUMNS
    )

    # title
    ws.merge_cells(
        f"A1:"
        f"{get_column_letter(num_cols)}1"
    )

    title = ws["A1"]

    title.value = (
        f"eCourts Legal Analytics Dashboard · "
        f"{user_name} · "
        f"Generated: "
        f"{datetime.now().strftime('%d %b %Y %H:%M')}"
    )

    title.font = Font(
        bold=True,
        size=13,
        color="FFFFFF",
    )

    title.fill = PatternFill(
        "solid",
        fgColor=C_TITLE_BG,
    )

    title.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.row_dimensions[
        1
    ].height = 30

    # headers
    h_fill = PatternFill(
        "solid",
        fgColor=C_HEADER_BG,
    )

    h_font = Font(
        bold=True,
        color=C_HEADER_FG,
    )

    for i, (
        label,
        _
    ) in enumerate(
        COLUMNS,
        start=1,
    ):

        c = ws.cell(
            row=2,
            column=i,
            value=label,
        )

        c.fill = h_fill
        c.font = h_font
        c.border = _thin_border()

        c.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        _set_col_width(
            ws,
            i,
            label,
        )

    ws.freeze_panes = "A3"

    alt_fill = PatternFill(
        "solid",
        fgColor=C_ALT_ROW,
    )

    fail_fill = PatternFill(
        "solid",
        fgColor=C_FAIL_ROW,
    )

    # rows
    for idx, result in enumerate(
        cases
    ):

        row = idx + 3

        fill = None

        if (
            result.get(
                "status"
            )
            == "failed"
        ):
            fill = fail_fill

        elif idx % 2:
            fill = alt_fill

        for col, (
            _,
            extractor,
        ) in enumerate(
            COLUMNS,
            start=1,
        ):

            try:
                value = (
                    extractor(
                        result
                    )
                )
            except Exception:
                value = ""

            cell = ws.cell(
                row=row,
                column=col,
                value=(
                    str(value)
                    if value != ""
                    else ""
                ),
            )

            cell.border = (
                _thin_border()
            )

            cell.font = Font(
                size=9
            )

            cell.alignment = (
                Alignment(
                    vertical="top",
                    wrap_text=True,
                )
            )

            if fill:
                cell.fill = fill

            _set_col_width(
                ws,
                col,
                value,
            )

    # summary
    gap = len(cases) + 4

    success = sum(
        1
        for c in cases
        if c.get(
            "status"
        ) == "done"
    )

    failed = (
        len(cases)
        - success
    )

    avg_time = round(
        sum(
            c.get(
                "processing_time",
                0,
            )
            for c in cases
        )
        / max(
            len(cases),
            1,
        ),
        2,
    )

    summary = [
        (
            "Total CNRs",
            len(cases),
        ),
        (
            "Successful",
            success,
        ),
        (
            "Failed",
            failed,
        ),
        (
            "Success %",
            round(
                success
                * 100
                / max(
                    len(cases),
                    1,
                ),
                2,
            ),
        ),
        (
            "Average Processing Time",
            f"{avg_time}s",
        ),
        (
            "Generated",
            datetime.now()
            .strftime(
                "%d %b %Y %H:%M"
            ),
        ),
    ]

    ws.cell(
        row=gap,
        column=1,
        value="── REPORT SUMMARY ──",
    ).font = Font(
        bold=True,
        size=11,
    )

    fill = PatternFill(
        "solid",
        fgColor=C_SUMMARY_BG,
    )

    for i, (
        k,
        v,
    ) in enumerate(
        summary
    ):

        r = gap + i + 1

        kc = ws.cell(
            r,
            1,
            k,
        )

        vc = ws.cell(
            r,
            2,
            str(v),
        )

        kc.fill = fill
        vc.fill = fill

        kc.border = (
            vc.border
        ) = _thin_border()

        kc.font = Font(
            bold=True
        )

    wb.save(
        filepath
    )

    print(
        f"[excel] "
        f"saved: "
        f"{filepath}"
    )

    return filepath